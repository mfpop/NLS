from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any

from data_management.models import ErpSourceFile
from data_management.services.erp_pattern_service import ErpPatternService


VALIDATION_STATUS_READY = "READY"
VALIDATION_STATUS_MISSING_FILE = "MISSING_FILE"
VALIDATION_STATUS_MISSING_FIELDS = "MISSING_FIELDS"
VALIDATION_STATUS_INVALID_FILE = "INVALID_FILE"
VALIDATION_STATUS_IMPORTED = "IMPORTED"
VALIDATION_STATUS_FAILED = "FAILED"


@dataclass
class ValidationResult:
    status: str
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    missing_fields: list[str] = field(default_factory=list)
    source_file: ErpSourceFile | None = None
    pattern: ErpPattern | None = None


class ErpImportValidationError(ValueError):
    def __init__(self, field: str, code: str, message: str):
        super().__init__(message)
        self.field = field
        self.code = code
        self.message = message


def _parse_file_headers(file_path: str) -> list[str]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None or ws.max_column is None or ws.max_column == 0:
                return []
            headers: list[str] = []
            for cell in ws[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
            wb.close()
            return headers
        except Exception:
            raise ErpImportValidationError("file", "UNREADABLE", "Cannot read Excel file headers")
    elif ext == ".csv":
        try:
            import csv
            with open(file_path, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                try:
                    row = next(reader)
                except StopIteration:
                    return []
                return [h.strip() for h in row if h.strip()]
        except Exception:
            raise ErpImportValidationError("file", "UNREADABLE", "Cannot read CSV file headers")
    elif ext == ".tsv":
        try:
            import csv
            with open(file_path, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter="\t")
                try:
                    row = next(reader)
                except StopIteration:
                    return []
                return [h.strip() for h in row if h.strip()]
        except Exception:
            raise ErpImportValidationError("file", "UNREADABLE", "Cannot read TSV file headers")
    elif ext == ".json":
        try:
            import json as json_lib
            with open(file_path, encoding="utf-8") as f:
                data = json_lib.load(f)
            if isinstance(data, list) and len(data) > 0:
                return list(data[0].keys()) if isinstance(data[0], dict) else []
            if isinstance(data, dict):
                records = data.get("records", data.get("data", []))
                if isinstance(records, list) and len(records) > 0:
                    return list(records[0].keys()) if isinstance(records[0], dict) else []
            return []
        except Exception:
            raise ErpImportValidationError("file", "UNREADABLE", "Cannot read JSON file headers")
    elif ext == ".xml":
        try:
            import xml.etree.ElementTree as ET
            tree = ET.parse(file_path)
            root = tree.getroot()
            first_child = None
            for child in root:
                first_child = child
                break
            if first_child is not None:
                return [child.tag for child in first_child]
            return []
        except Exception:
            raise ErpImportValidationError("file", "UNREADABLE", "Cannot read XML file headers")
    return []


def _find_source_file_for_pattern(pattern) -> ErpSourceFile | None:
    """Match a source file to a pattern using file type and optional glob pattern.

    Accepts both manufacturing.ErpImportPattern (has source_file_pattern)
    and legacy data_management.ErpPattern (has source_file_type only).
    """
    file_type = getattr(pattern, "source_file_type", None) or ""
    glob_pattern = getattr(pattern, "source_file_pattern", None) or ""
    if not file_type and glob_pattern:
        file_type = os.path.splitext(glob_pattern)[1].lstrip(".").lower()
    if not file_type:
        return None
    candidates = ErpSourceFile.objects.filter(
        file_type=file_type,
        status__in=("UPLOADED", "VALIDATED"),
        original_name__endswith=f".{file_type}",
    ).order_by("-uploaded_at")
    if glob_pattern:
        import fnmatch
        for sf in candidates:
            if fnmatch.fnmatch(sf.original_name, glob_pattern):
                return sf
        return None
    return candidates.first()


class ErpImportValidationService:

    @staticmethod
    def validate_source_file_exists(pattern) -> ErpSourceFile | None:
        return _find_source_file_for_pattern(pattern)

    @staticmethod
    def validate_file_type_matches(pattern, source_file: ErpSourceFile) -> bool:
        file_type = getattr(pattern, "source_file_type", None) or ""
        return file_type == source_file.file_type

    @staticmethod
    def parse_file_headers(source_file: ErpSourceFile) -> list[str]:
        if not os.path.isfile(source_file.file_path):
            raise ErpImportValidationError("file", "NOT_FOUND", f"Source file not found at {source_file.file_path}")
        return _parse_file_headers(source_file.file_path)

    @staticmethod
    def validate_mapped_source_fields(pattern, headers: list[str]) -> list[str]:
        header_set = {h.lower().strip() for h in headers}
        missing: list[str] = []
        for mapping in pattern.field_mappings.all().order_by("order"):
            if mapping.source_name.lower().strip() not in header_set:
                missing.append(mapping.source_name)
        return missing

    @staticmethod
    def validate_required_mappings_complete(pattern) -> list[str]:
        missing: list[str] = []
        for mapping in pattern.field_mappings.all().order_by("order"):
            if mapping.is_required and not mapping.source_name.strip():
                missing.append(mapping.destination_name)
        return missing

    @staticmethod
    def validate_destination_entity(pattern) -> list[str]:
        from data_management.services.erp_pattern_service import (
            ErpPatternService, VALID_DESTINATION_ENTITIES,
        )
        errors: list[str] = []
        if pattern.destination_entity not in VALID_DESTINATION_ENTITIES:
            errors.append(f"Unsupported destination entity '{pattern.destination_entity}'")
        return errors

    @staticmethod
    def validate_pattern(pattern_id: int) -> ValidationResult:
        try:
            pattern = ErpPatternService.get_pattern(pattern_id)
        except Exception as e:
            return ValidationResult(
                status=VALIDATION_STATUS_FAILED,
                errors=[str(e)],
            )

        errors: list[str] = []
        warnings: list[str] = []

        source_file = ErpImportValidationService.validate_source_file_exists(pattern)
        if source_file is None:
            return ValidationResult(
                status=VALIDATION_STATUS_MISSING_FILE,
                errors=["No matching source file found for this pattern"],
                pattern=pattern,
            )

        if not ErpImportValidationService.validate_file_type_matches(pattern, source_file):
            return ValidationResult(
                status=VALIDATION_STATUS_INVALID_FILE,
                errors=[
                    f"File type mismatch: pattern expects '{pattern.source_file_type}' "
                    f"but source file is '{source_file.file_type}'"
                ],
                source_file=source_file,
                pattern=pattern,
            )

        try:
            headers = ErpImportValidationService.parse_file_headers(source_file)
        except ErpImportValidationError as e:
            return ValidationResult(
                status=VALIDATION_STATUS_INVALID_FILE,
                errors=[str(e)],
                source_file=source_file,
                pattern=pattern,
            )

        if not headers:
            return ValidationResult(
                status=VALIDATION_STATUS_INVALID_FILE,
                errors=["File has no readable headers"],
                source_file=source_file,
                pattern=pattern,
            )

        dest_errors = ErpImportValidationService.validate_destination_entity(pattern)
        if dest_errors:
            errors.extend(dest_errors)

        missing_fields = ErpImportValidationService.validate_mapped_source_fields(pattern, headers)
        if missing_fields:
            errors.append(f"Missing mapped fields in source file: {', '.join(missing_fields)}")
            return ValidationResult(
                status=VALIDATION_STATUS_MISSING_FIELDS,
                errors=errors,
                warnings=warnings,
                missing_fields=missing_fields,
                source_file=source_file,
                pattern=pattern,
            )

        required_missing = ErpImportValidationService.validate_required_mappings_complete(pattern)
        if required_missing:
            errors.append(f"Required mappings incomplete: {', '.join(required_missing)}")
            return ValidationResult(
                status=VALIDATION_STATUS_MISSING_FIELDS,
                errors=errors,
                warnings=warnings,
                missing_fields=required_missing,
                source_file=source_file,
                pattern=pattern,
            )

        if errors:
            return ValidationResult(
                status=VALIDATION_STATUS_FAILED,
                errors=errors,
                warnings=warnings,
                source_file=source_file,
                pattern=pattern,
            )

        return ValidationResult(
            status=VALIDATION_STATUS_READY,
            errors=[],
            warnings=warnings,
            source_file=source_file,
            pattern=pattern,
        )
