from __future__ import annotations
import logging
import os
from pathlib import Path
from typing import Any
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from django.db import transaction
from django.core.files.storage import default_storage

from manufacturing.models import (
    Company, Plant, ProductionLine, Department,
    ProductionLineDepartmentAssignment,
    ResourceGroup, Resource, EntityStatus,
)

logger = logging.getLogger(__name__)

REQUIRED_SHEETS = [
    "Companies", "Plants", "ProductionLines", "Departments",
    "ProductionLineDepartments", "ResourceGroups", "Resources",
]

SHEET_COLUMNS = {
    "Companies": ["company_code", "company_name", "status"],
    "Plants": ["company_code", "plant_code", "plant_name", "status"],
    "ProductionLines": ["plant_code", "line_code", "line_name", "status"],
    "Departments": ["plant_code", "department_code", "department_name", "status"],
    "ProductionLineDepartments": ["plant_code", "line_code", "department_code", "sequence", "status"],
    "ResourceGroups": ["plant_code", "department_code", "resource_group_code", "resource_group_name", "status"],
    "Resources": ["plant_code", "department_code", "resource_group_code", "resource_code", "resource_name", "status"],
}

VALID_STATUSES = {s.value for s in EntityStatus}


@dataclass
class ParsedRow:
    sheet: str
    row_number: int
    data: dict[str, str]


@dataclass
class ParsedWorkbook:
    sheets: dict[str, list[ParsedRow]] = field(default_factory=dict)


@dataclass
class ValidationError:
    sheet: str
    row_number: int
    field: str
    message: str


@dataclass
class ImportResult:
    ok: bool
    validation_errors: list[ValidationError] = field(default_factory=list)
    companies_created: int = 0
    companies_updated: int = 0
    plants_created: int = 0
    plants_updated: int = 0
    lines_created: int = 0
    lines_updated: int = 0
    departments_created: int = 0
    departments_updated: int = 0
    assignments_created: int = 0
    assignments_updated: int = 0
    resource_groups_created: int = 0
    resource_groups_updated: int = 0
    resources_created: int = 0
    resources_updated: int = 0


class ImportMode:
    VALIDATE_ONLY = "VALIDATE_ONLY"
    COMPARE_ONLY = "COMPARE_ONLY"
    UPSERT = "UPSERT"
    REJECT_ON_DIFF = "REJECT_ON_DIFF"


class PlantStructureImportService:

    @staticmethod
    def parse_excel(file_path: str) -> ParsedWorkbook:
        if os.path.exists(file_path):
            abs_path = file_path
        elif default_storage.exists(file_path):
            abs_path = default_storage.path(file_path)
        else:
            abs_path = file_path
        wb = load_workbook(abs_path, read_only=True, data_only=True)
        result = ParsedWorkbook()
        for sheet_name in REQUIRED_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            header = [str(c).strip().lower() if c else "" for c in next(rows, [])]
            col_map = {col: idx for idx, col in enumerate(header)}
            parsed: list[ParsedRow] = []
            for i, row in enumerate(rows, start=2):
                row_data = {}
                for col_name, col_idx in col_map.items():
                    if col_idx < len(row):
                        val = row[col_idx]
                        row_data[col_name] = str(val).strip() if val is not None else ""
                if any(row_data.values()):
                    parsed.append(ParsedRow(sheet=sheet_name, row_number=i, data=row_data))
            result.sheets[sheet_name] = parsed
        wb.close()
        return result

    @staticmethod
    def validate_workbook(parsed: ParsedWorkbook) -> list[ValidationError]:
        errors: list[ValidationError] = []
        for sheet_name in REQUIRED_SHEETS:
            if sheet_name not in parsed.sheets:
                errors.append(ValidationError(sheet=sheet_name, row_number=0, field="sheet", message=f"Missing required sheet: {sheet_name}"))
        return errors

    @staticmethod
    def validate_rows(parsed: ParsedWorkbook) -> list[ValidationError]:
        errors: list[ValidationError] = []
        for sheet_name, cols in SHEET_COLUMNS.items():
            for row in parsed.sheets.get(sheet_name, []):
                for col in cols:
                    if col.endswith("_code") or col.endswith("_name"):
                        if not row.data.get(col, "").strip():
                            errors.append(ValidationError(sheet=sheet_name, row_number=row.row_number, field=col, message=f"Required field '{col}' is empty"))
                status_val = row.data.get("status", "").upper()
                if status_val and status_val not in VALID_STATUSES:
                    errors.append(ValidationError(sheet=sheet_name, row_number=row.row_number, field="status", message=f"Invalid status '{status_val}'. Must be one of: {', '.join(sorted(VALID_STATUSES))}"))

        seen: dict[str, set[str]] = {}
        for sheet_name, cols in SHEET_COLUMNS.items():
            key_fields = [c for c in cols if c.endswith("_code") and not c.endswith("_name")]
            for kf in key_fields:
                seen_key = f"{sheet_name}:{kf}"
                seen_values: set[str] = set()
                for row in parsed.sheets.get(sheet_name, []):
                    val = row.data.get(kf, "").strip()
                    if not val:
                        continue
                    if val in seen_values:
                        errors.append(ValidationError(sheet=sheet_name, row_number=row.row_number, field=kf, message=f"Duplicate {kf} '{val}' in Excel"))
                    seen_values.add(val)
                seen[seen_key] = seen_values

        return errors

    @staticmethod
    def _validate_hierarchy(parsed: ParsedWorkbook) -> list[ValidationError]:
        errors: list[ValidationError] = []

        companies = {r.data.get("company_code", "").strip() for r in parsed.sheets.get("Companies", [])}
        plants = {r.data.get("plant_code", "").strip() for r in parsed.sheets.get("Plants", [])}
        plant_to_company = {}
        for r in parsed.sheets.get("Plants", []):
            pc = r.data.get("plant_code", "").strip()
            cc = r.data.get("company_code", "").strip()
            if pc and cc:
                plant_to_company[pc] = cc

        departments = {r.data.get("department_code", "").strip() for r in parsed.sheets.get("Departments", [])}
        dept_to_plant = {}
        for r in parsed.sheets.get("Departments", []):
            dc = r.data.get("department_code", "").strip()
            pc = r.data.get("plant_code", "").strip()
            if dc and pc:
                dept_to_plant[dc] = pc

        lines = {r.data.get("line_code", "").strip() for r in parsed.sheets.get("ProductionLines", [])}
        line_to_plant = {}
        for r in parsed.sheets.get("ProductionLines", []):
            lc = r.data.get("line_code", "").strip()
            pc = r.data.get("plant_code", "").strip()
            if lc and pc:
                line_to_plant[lc] = pc

        resource_groups = {r.data.get("resource_group_code", "").strip() for r in parsed.sheets.get("ResourceGroups", [])}
        rg_to_dept = {}
        for r in parsed.sheets.get("ResourceGroups", []):
            rc = r.data.get("resource_group_code", "").strip()
            dc = r.data.get("department_code", "").strip()
            if rc and dc:
                rg_to_dept[rc] = dc

        rgs_in_dept_plant = {}
        for rc, dc in rg_to_dept.items():
            rgs_in_dept_plant[rc] = dept_to_plant.get(dc)

        for row in parsed.sheets.get("Plants", []):
            cc = row.data.get("company_code", "").strip()
            if cc and cc not in companies:
                errors.append(ValidationError(sheet="Plants", row_number=row.row_number, field="company_code", message=f"Company '{cc}' not found in Companies sheet"))

        for row in parsed.sheets.get("ProductionLines", []):
            pc = row.data.get("plant_code", "").strip()
            if pc and pc not in plants:
                errors.append(ValidationError(sheet="ProductionLines", row_number=row.row_number, field="plant_code", message=f"Plant '{pc}' not found in Plants sheet"))

        for row in parsed.sheets.get("Departments", []):
            pc = row.data.get("plant_code", "").strip()
            if pc and pc not in plants:
                errors.append(ValidationError(sheet="Departments", row_number=row.row_number, field="plant_code", message=f"Plant '{pc}' not found in Plants sheet"))

        for row in parsed.sheets.get("ProductionLineDepartments", []):
            pc = row.data.get("plant_code", "").strip()
            lc = row.data.get("line_code", "").strip()
            dc = row.data.get("department_code", "").strip()

            if pc and pc not in plants:
                errors.append(ValidationError(sheet="ProductionLineDepartments", row_number=row.row_number, field="plant_code", message=f"Plant '{pc}' not found in Plants sheet"))
            if lc and lc not in lines:
                errors.append(ValidationError(sheet="ProductionLineDepartments", row_number=row.row_number, field="line_code", message=f"Line '{lc}' not found in ProductionLines sheet"))
            if dc and dc not in departments:
                errors.append(ValidationError(sheet="ProductionLineDepartments", row_number=row.row_number, field="department_code", message=f"Department '{dc}' not found in Departments sheet"))
            if lc and dc and line_to_plant.get(lc) != dept_to_plant.get(dc):
                errors.append(ValidationError(sheet="ProductionLineDepartments", row_number=row.row_number, field="line_code", message=f"Line '{lc}' (plant {line_to_plant.get(lc)}) and Department '{dc}' (plant {dept_to_plant.get(dc)}) belong to different plants"))

        for row in parsed.sheets.get("ResourceGroups", []):
            pc = row.data.get("plant_code", "").strip()
            dc = row.data.get("department_code", "").strip()

            if pc and pc not in plants:
                errors.append(ValidationError(sheet="ResourceGroups", row_number=row.row_number, field="plant_code", message=f"Plant '{pc}' not found in Plants sheet"))
            if dc and dc not in departments:
                errors.append(ValidationError(sheet="ResourceGroups", row_number=row.row_number, field="department_code", message=f"Department '{dc}' not found in Departments sheet"))
            if dc and dept_to_plant.get(dc) and dept_to_plant.get(dc) != pc:
                errors.append(ValidationError(sheet="ResourceGroups", row_number=row.row_number, field="department_code", message=f"Department '{dc}' belongs to plant {dept_to_plant.get(dc)}, not {pc}"))

        for row in parsed.sheets.get("Resources", []):
            pc = row.data.get("plant_code", "").strip()
            dc = row.data.get("department_code", "").strip()
            rc = row.data.get("resource_group_code", "").strip()

            if pc and pc not in plants:
                errors.append(ValidationError(sheet="Resources", row_number=row.row_number, field="plant_code", message=f"Plant '{pc}' not found in Plants sheet"))
            if dc and dc not in departments:
                errors.append(ValidationError(sheet="Resources", row_number=row.row_number, field="department_code", message=f"Department '{dc}' not found in Departments sheet"))
            if rc and rc not in resource_groups:
                errors.append(ValidationError(sheet="Resources", row_number=row.row_number, field="resource_group_code", message=f"Resource group '{rc}' not found in ResourceGroups sheet"))
            if rc and rgs_in_dept_plant.get(rc) and dept_to_plant.get(dc) and rgs_in_dept_plant[rc] != dept_to_plant.get(dc):
                errors.append(ValidationError(sheet="Resources", row_number=row.row_number, field="resource_group_code", message=f"Resource group '{rc}' belongs to plant {rgs_in_dept_plant[rc]}, resource is assigned to plant {dept_to_plant.get(dc)}"))

        return errors

    @staticmethod
    def import_workbook(file_path: str, mode: str) -> ImportResult:
        parsed = PlantStructureImportService.parse_excel(file_path)
        result = ImportResult(ok=True)

        ws_errors = PlantStructureImportService.validate_workbook(parsed)
        if ws_errors:
            result.ok = False
            result.validation_errors.extend(ws_errors)
            return result

        row_errors = PlantStructureImportService.validate_rows(parsed)
        if row_errors:
            result.ok = False
            result.validation_errors.extend(row_errors)

        hierarchy_errors = PlantStructureImportService._validate_hierarchy(parsed)
        if hierarchy_errors:
            result.ok = False
            result.validation_errors.extend(hierarchy_errors)

        if not result.ok or mode == ImportMode.VALIDATE_ONLY or mode == ImportMode.COMPARE_ONLY:
            return result

        if mode == ImportMode.UPSERT:
            PlantStructureImportService._upsert_all(parsed, result)

        return result

    @staticmethod
    @transaction.atomic
    def _upsert_all(parsed: ParsedWorkbook, result: ImportResult) -> None:
        company_map: dict[str, Company] = {}
        for row in parsed.sheets.get("Companies", []):
            d = row.data
            code = d.get("company_code", "").strip()
            if not code:
                continue
            company, created = Company.objects.update_or_create(
                code=code,
                defaults={
                    "name": d.get("company_name", "").strip() or code,
                    "status": d.get("status", "").upper() or EntityStatus.ACTIVE,
                },
            )
            company_map[code] = company
            if created:
                result.companies_created += 1
            else:
                result.companies_updated += 1

        plant_map: dict[str, Plant] = {}
        for row in parsed.sheets.get("Plants", []):
            d = row.data
            code = d.get("plant_code", "").strip()
            cc = d.get("company_code", "").strip()
            if not code or not cc:
                continue
            company = company_map.get(cc)
            if not company:
                continue
            plant, created = Plant.objects.update_or_create(
                code=code, company=company,
                defaults={
                    "name": d.get("plant_name", "").strip() or code,
                    "status": d.get("status", "").upper() or EntityStatus.ACTIVE,
                },
            )
            plant_map[code] = plant
            if created:
                result.plants_created += 1
            else:
                result.plants_updated += 1

        line_map: dict[str, ProductionLine] = {}
        for row in parsed.sheets.get("ProductionLines", []):
            d = row.data
            code = d.get("line_code", "").strip()
            pc = d.get("plant_code", "").strip()
            if not code or not pc:
                continue
            plant = plant_map.get(pc)
            if not plant:
                continue
            line_obj, created = ProductionLine.objects.update_or_create(
                code=code, plant=plant,
                defaults={
                    "name": d.get("line_name", "").strip() or code,
                    "status": d.get("status", "").upper() or EntityStatus.ACTIVE,
                },
            )
            line_map[code] = line_obj
            if created:
                result.lines_created += 1
            else:
                result.lines_updated += 1

        dept_map: dict[str, Department] = {}
        for row in parsed.sheets.get("Departments", []):
            d = row.data
            code = d.get("department_code", "").strip()
            pc = d.get("plant_code", "").strip()
            if not code or not pc:
                continue
            plant = plant_map.get(pc)
            if not plant:
                continue
            dept, created = Department.objects.update_or_create(
                code=code, plant=plant,
                defaults={
                    "name": d.get("department_name", "").strip() or code,
                    "status": d.get("status", "").upper() or EntityStatus.ACTIVE,
                },
            )
            dept_map[code] = dept
            if created:
                result.departments_created += 1
            else:
                result.departments_updated += 1

        for row in parsed.sheets.get("ProductionLineDepartments", []):
            d = row.data
            lc = d.get("line_code", "").strip()
            dc = d.get("department_code", "").strip()
            pc = d.get("plant_code", "").strip()
            seq_str = d.get("sequence", "").strip()
            sequence = int(seq_str) if seq_str.isdigit() else 0
            if not lc or not dc or not pc:
                continue
            plant = plant_map.get(pc)
            line_obj = line_map.get(lc)
            dept = dept_map.get(dc)
            if not plant or not line_obj or not dept:
                continue
            _, created = ProductionLineDepartmentAssignment.objects.update_or_create(
                production_line=line_obj, department=dept,
                defaults={
                    "plant": plant,
                    "sequence": sequence,
                    "status": d.get("status", "").upper() or EntityStatus.ACTIVE,
                },
            )
            if created:
                result.assignments_created += 1
            else:
                result.assignments_updated += 1

        rg_map: dict[str, ResourceGroup] = {}
        for row in parsed.sheets.get("ResourceGroups", []):
            d = row.data
            code = d.get("resource_group_code", "").strip()
            dc = d.get("department_code", "").strip()
            if not code or not dc:
                continue
            dept = dept_map.get(dc)
            if not dept:
                continue
            rg, created = ResourceGroup.objects.update_or_create(
                code=code, department=dept,
                defaults={
                    "name": d.get("resource_group_name", "").strip() or code,
                    "status": d.get("status", "").upper() or EntityStatus.ACTIVE,
                },
            )
            rg_map[code] = rg
            if created:
                result.resource_groups_created += 1
            else:
                result.resource_groups_updated += 1

        for row in parsed.sheets.get("Resources", []):
            d = row.data
            code = d.get("resource_code", "").strip()
            rc = d.get("resource_group_code", "").strip()
            if not code or not rc:
                continue
            rg = rg_map.get(rc)
            if not rg:
                continue
            _, created = Resource.objects.update_or_create(
                code=code, resource_group=rg,
                defaults={
                    "name": d.get("resource_name", "").strip() or code,
                    "status": d.get("status", "").upper() or EntityStatus.ACTIVE,
                },
            )
            if created:
                result.resources_created += 1
            else:
                result.resources_updated += 1
