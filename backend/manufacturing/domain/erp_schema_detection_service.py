from __future__ import annotations

import json as json_lib
import os
from dataclasses import dataclass, field
from typing import Any

from data_management.models import ErpSourceFile


@dataclass
class SchemaDetectionResult:
    file_type: str
    sheet_names: list[str] = field(default_factory=list)
    selected_sheet_name: str | None = None
    source_schema: list[dict[str, str]] = field(default_factory=list)
    requires_sheet_selection: bool = False


class SchemaDetectionError(ValueError):
    def __init__(self, field: str, code: str, message: str):
        super().__init__(message)
        self.field = field
        self.code = code
        self.message = message


def _detect_excel_sheet_names(file_path: str) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _parse_excel_sheet_schema(file_path: str, sheet_name: str) -> list[dict[str, str]]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SchemaDetectionError("selectedSheetName", "INVALID_FILE", f"Sheet '{sheet_name}' not found in workbook")
    ws = wb[sheet_name]
    schema: list[dict[str, str]] = []
    for cell in ws[1]:
        if cell.value is not None:
            header = str(cell.value).strip()
            if header:
                schema.append({"fieldName": header, "dataType": "string"})
    wb.close()
    return schema


def _detect_csv_schema(file_path: str) -> list[dict[str, str]]:
    import csv
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        try:
            row = next(reader)
        except StopIteration:
            return []
        return [{"fieldName": h.strip(), "dataType": "string"} for h in row if h.strip()]


def _detect_json_schema(file_path: str) -> list[dict[str, str]]:
    with open(file_path, encoding="utf-8") as f:
        data = json_lib.load(f)
    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
        return [{"fieldName": k, "dataType": "string"} for k in data[0].keys()]
    if isinstance(data, dict):
        records = data.get("records", data.get("data", []))
        if isinstance(records, list) and len(records) > 0 and isinstance(records[0], dict):
            return [{"fieldName": k, "dataType": "string"} for k in records[0].keys()]
    return []


def _detect_xml_schema(file_path: str) -> list[dict[str, str]]:
    import xml.etree.ElementTree as ET
    tree = ET.parse(file_path)
    root = tree.getroot()
    first_child = None
    for child in root:
        first_child = child
        break
    if first_child is not None:
        return [{"fieldName": child.tag, "dataType": "string"} for child in first_child]
    return []


def _detect_tsv_schema(file_path: str) -> list[dict[str, str]]:
    import csv
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter="\t")
        try:
            row = next(reader)
        except StopIteration:
            return []
        return [{"fieldName": h.strip(), "dataType": "string"} for h in row if h.strip()]


class ErpSchemaDetectionService:

    @staticmethod
    def detect_schema(
        source_file: ErpSourceFile,
        selected_sheet_name: str | None = None,
    ) -> SchemaDetectionResult:
        if not os.path.isfile(source_file.file_path):
            raise SchemaDetectionError("file", "NOT_FOUND", "Source file not found")

        ext = source_file.file_type.lower()

        if ext in ("xlsx", "xls"):
            sheet_names = _detect_excel_sheet_names(source_file.file_path)
            if len(sheet_names) == 1:
                selected = sheet_names[0]
                schema = _parse_excel_sheet_schema(source_file.file_path, selected)
                return SchemaDetectionResult(
                    file_type=ext,
                    sheet_names=sheet_names,
                    selected_sheet_name=selected,
                    source_schema=schema,
                    requires_sheet_selection=False,
                )
            elif len(sheet_names) > 1:
                if selected_sheet_name:
                    if selected_sheet_name not in sheet_names:
                        raise SchemaDetectionError(
                            "selectedSheetName", "INVALID_FILE",
                            f"Selected sheet '{selected_sheet_name}' not found in workbook. Available: {', '.join(sheet_names)}",
                        )
                    schema = _parse_excel_sheet_schema(source_file.file_path, selected_sheet_name)
                    return SchemaDetectionResult(
                        file_type=ext,
                        sheet_names=sheet_names,
                        selected_sheet_name=selected_sheet_name,
                        source_schema=schema,
                        requires_sheet_selection=False,
                    )
                return SchemaDetectionResult(
                    file_type=ext,
                    sheet_names=sheet_names,
                    selected_sheet_name=None,
                    source_schema=[],
                    requires_sheet_selection=True,
                )
            return SchemaDetectionResult(
                file_type=ext,
                sheet_names=[],
                selected_sheet_name=None,
                source_schema=[],
                requires_sheet_selection=False,
            )

        ext_map = {
            "csv": _detect_csv_schema,
            "tsv": _detect_tsv_schema,
            "json": _detect_json_schema,
            "xml": _detect_xml_schema,
        }
        detector = ext_map.get(ext)
        if detector:
            schema = detector(source_file.file_path)
            return SchemaDetectionResult(
                file_type=ext,
                sheet_names=[],
                selected_sheet_name=None,
                source_schema=schema,
                requires_sheet_selection=False,
            )

        raise SchemaDetectionError("file", "UNSUPPORTED", f"Unsupported file type: {ext}")
